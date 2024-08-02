from django.shortcuts import render, redirect,redirect,get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from django.contrib import messages
from django.urls import reverse 
from django.contrib.auth.decorators import login_required
from curriculum.models import *
from users.models import *
from curriculum import views
from django.core.cache import cache
from django.contrib.admin.models import LogEntry
import datetime as dt
from .signals import succesful_logout
from assessment.models import *
from django.db.models import Q,Max
from django.core.exceptions import ObjectDoesNotExist
from users.forms import *
import openpyxl
import logging
from django.db.models import Count
from collections import defaultdict
from django.db.models import OuterRef, Subquery

def teacher_home(request,subject_id=None):
    teacher_obj = user_profile_teacher.objects.get(user=request.user)
    
    school=teacher_obj.school
    students=user_profile_student.objects.filter(school__icontains=school).count()

    subjects=Subject.objects.all().count()
    advocacy_count=AdvocacyVisit.objects.filter(school__icontains=school).count()
    advocacy=AdvocacyVisit.objects.filter(school__icontains=school)

    unread_notifications = NotificationTeacher.objects.filter(teacher_id=teacher_obj, read=False).count()
    
    allowed_emails = [
        "training1@thinnkware.com", "training@thinnkware.com", "training2@thinnkware.com", 
        "training3@thinnkware.com", "training4@thinnkware.com", "training5@thinnkware.com", 
        "training6@thinnkware.com", "training7@thinnkware.com", "training8@thinnkware.com", 
        "training9@thinnkware.com", "training10@thinnkware.com", "training11@thinnkware.com", 
        "training12@thinnkware.com", "training13@thinnkware.com",
        "manager.lnd@thinnkware.com","manager.lnd@thinnkware.com",
        "training01@thinnkware.com",
        "msraafas30@gmail.com","training14@thinnkware.com","techsupport@thinnkware.com","devesh.d1560@gmail.com",
        "developer2@thinnkware.com"
    ]
    
    rofl_access=["teacher_rofl@thinnkware.com"]
    st_mary=["do.amalan@gmail.com","teacher_roflschool@thinnkware.com"]
    
    teacher_sheet_urls = {
    "training01@thinnkware.com": "https://docs.google.com/spreadsheets/d/1yqQL_dFTKq5t9M4sdSuOrXQbekteDgIw/edit",
    "training1@thinnkware.com": "https://docs.google.com/spreadsheets/d/1yqQL_dFTKq5t9M4sdSuOrXQbekteDgIw/edit",
    "training2@thinnkware.com": "https://docs.google.com/spreadsheets/d/1cbHHCc-sf5ctuPfk5wQ6LuJEL_b7Q6Tj/edit",
    "training3@thinnkware.com": "https://docs.google.com/spreadsheets/d/1cbHHCc-sf5ctuPfk5wQ6LuJEL_b7Q6Tj/edit",
    "training4@thinnkware.com": "https://docs.google.com/spreadsheets/d/1PJmXIgu8v_V10FIKL3AW7teGu8XcAlqkIc8YeDKzaOo/edit",
    "training5@thinnkware.com": "https://docs.google.com/spreadsheets/d/1EnxRAMA39cXEJ1_BDcevhkUf42tUyxQH/edit",
    "training6@thinnkware.com": "https://docs.google.com/spreadsheets/d/1yqQL_dFTKq5t9M4sdSuOrXQbekteDgIw/edit",
    "training7@thinnkware.com": "https://docs.google.com/spreadsheets/d/1EnxRAMA39cXEJ1_BDcevhkUf42tUyxQH/edit",
    "training8@thinnkware.com": "https://docs.google.com/spreadsheets/d/1HSCRazedpWJ9Hc7y1womiB4JBKtW3SJYA1S4FILN_4A/edit",
    "training9@thinnkware.com": "https://docs.google.com/spreadsheets/d/16X2VQWQIeGnXsjRTlz9AVLYEZDrkeRnb/edit",
    "training10@thinnkware.com": "https://docs.google.com/spreadsheets/d/1PJmXIgu8v_V10FIKL3AW7teGu8XcAlqkIc8YeDKzaOo/edit",
    "training11@thinnkware.com": "https://docs.google.com/spreadsheets/d/1HSCRazedpWJ9Hc7y1womiB4JBKtW3SJYA1S4FILN_4A/edit",
    "training12@thinnkware.com": "https://docs.google.com/spreadsheets/d/15v4ENhyBS2uq4cc0D-4WTir5aUcuDWedCqyjcz87tt0/edit",
    "training13@thinnkware.com": "https://docs.google.com/spreadsheets/d/1sw5RveApNz_8IfWaEOd5qkf8EXYgRd6uONfjfZCXKUI/edit",
    "training14@thinnkware.com": "https://docs.google.com/spreadsheets/d//edit",
    "training15@thinnkware.com": "https://docs.google.com/spreadsheets/d//edit",}

    teacher_sheet_url = teacher_sheet_urls.get(request.user.email, "")
    gallery_school=SchoolGallery.objects.filter(school__icontains=school).first()
    observation_sheet=ObservationSheet.objects.filter(school__icontains=school).first()
    
    if request.user.email in allowed_emails:
        context={
            "total_student": students,
            "total_advocacy":advocacy_count,
            "advocacy":advocacy,
            "school":teacher_obj,
            "unread_notifications":unread_notifications,
            "subjects":subjects,
            "allowed_emails":allowed_emails,
            "teacher_sheet_url":teacher_sheet_url,
            "gallery_school":gallery_school,
            "observation_sheet":observation_sheet,
        }
    else:
        context={
            "total_student": students,
            "school":teacher_obj,
            "unread_notifications":unread_notifications,
            "subjects":subjects,
            "rofl_access":rofl_access,
            "st_mary":st_mary
        }
    return render(request, "teacher_template/teacher_home_template.html", context)


def subject_list(request):
    subjects={
        "Junior Scratch":"1-2",
        "Hands-On Science Kit":"1-8",
        "Scratch":"3",
        "Robotics": "3-9",
        "MIT App Inventor":"6",
        "3D Designing":"7",
        "Arduino": "7-10",
        "Python": "7-8",
        "Internet of Things":"9",
        "AI": "4-10",
        "Jodo Straw":"3-4",
        "Paper Circuit":"2-3",
        "Data Science":"9-10",
        "Early Simple Machine":"1-2",
        "Tinkercad":"4-6"
    }

    processed_subjects = {}
    for subject, classes in subjects.items():
        class_range = classes.split('-')
        processed_subjects[subject] = {
            "full_range": classes,
            "first_class": class_range[0]
        }
    
    return render(request, "teacher_template/subjects_list.html", {"subjects": processed_subjects})


def teacher_feedback(request):
    teacher_obj = user_profile_teacher.objects.get(user=request.user)
    feedback_data = FeedBackTeacher.objects.filter(teacher=teacher_obj)
    context = {
        "feedback_data": feedback_data,
        "school":teacher_obj
    }
    return render(request, 'teacher_template/teacher_feedback.html', context)

def teacher_feedback_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method.")
        return redirect('users:teacher_feedback')
    else:
        feedback = request.POST.get('feedback_message')
        teacher_obj = user_profile_teacher.objects.get(user=request.user)

        try:
            add_feedback = FeedBackTeacher(teacher=teacher_obj, feedback=feedback, feedback_reply="")
            add_feedback.save()
            # messages.success(request, "Feedback Sent.")
            return redirect('users:teacher_feedback')
        except:
            messages.error(request, "Failed to Send Feedback.")
            return redirect('users:teacher_feedback')
        
def notifications(request):
    user=request.user
    teacher=user_profile_teacher.objects.get(user=user)
    notifications = NotificationTeacher.objects.filter(teacher_id=teacher).order_by('-id')
    return render(request, 'teacher_template/notification.html', {'notifications': notifications,"school":teacher})

def mark_notification_as_read(request, id):
    notification = NotificationTeacher.objects.get(id=id)
    notification.read = True
    notification.save()
    return redirect('users:teacher_feedback')


def leaderboard(request):
    teacher=user_profile_teacher.objects.get(user=request.user)
    school=teacher.school
    grade=teacher.grade
    
    # max_results = Result.objects.values('user__user_profile_student__grade').annotate(max_score=Max('score'))
    overall_students = Result.objects.filter(user__user_profile_student__school__icontains=school).order_by('-score')[:5]
    
    # Fetch the overall top students from the teacher's grade
    grade_students = Result.objects.filter(
        user__user_profile_student__grade=grade,
        user__user_profile_student__school=school
    ).order_by('-score')[:3]

    # Initialize a dictionary to store top 3 students for each grade
    top_students_by_grade = {}
    
    max_results = Result.objects.filter(
        user__user_profile_student__school__icontains=school
    ).values('user__user_profile_student__grade').annotate(max_score=Max('score'))

    # Iterate through each grade and get the top 3 students
    for grade_info in max_results:
        grade = grade_info['user__user_profile_student__grade']
        max_score = grade_info['max_score']
        
        # Get the top 3 students with the maximum score in each grade
        top_students = Result.objects.filter(
            user__user_profile_student__grade=grade,
            user__user_profile_student__school__icontains=school
        ).order_by('-score')[:3]

        # Add the top students to the dictionary
        top_students_by_grade[grade] = top_students

    return render(request, 'teacher_template/leadership.html', {'top_students_by_grade': grade_students,"overall_students":overall_students,"school":teacher})
    
# def student_report(request):
#     try:
#         teacher = user_profile_teacher.objects.get(user=request.user)
#         school_name = teacher.school
#         students = user_profile_student.objects.filter(school=school_name)  # Use exact match for school
#     except user_profile_teacher.DoesNotExist:
#         return render(request, 'error_template.html', {'message': 'Teacher profile not found'})

#     students_gradewise = []

#     school_grade_mapping = {
#         "Sparsh Global School,Greater Noida": range(1, 10),
#         "JP International School,Greater Noida": range(1, 10),
#         "SPS,Sonipat": range(1, 10),
#         "Vivekanand School,Anand Vihar": range(6, 10),
#         "Blooming Dale School,Budaun": range(3, 9),
#         "Satya Prakash Public School, Jabalpur":range(1,10)
#     }

#     valid_grades = school_grade_mapping.get(school_name, None)

#     if valid_grades is None:
#         valid_grades = students.values_list('grade', flat=True).distinct()

#     for grade in valid_grades:
#         students_for_grade = students.filter(grade=grade)
#         total_number = students_for_grade.count()
#         students_gradewise.append((grade, students_for_grade, total_number))

#     search_query = request.GET.get('search', '')
#     if search_query:
#         students = students.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query))
#         students_gradewise = [(student.grade, [student], 1) for student in students]

#     context = {
#         "students_gradewise": students_gradewise,
#         "students": students,
#         "search_query": search_query,
#         "school": school_name
#     }

#     if request.is_ajax():
#         students_data = [{'grade': student.grade, 'student_name': student.first_name, 'student_id': student.user_id} for student in students]
#         return JsonResponse({'students': students_data})

#     return render(request, 'teacher_template/view_report.html', context)


# def student_report_gradewise(request,grade):
#     teacher=user_profile_teacher.objects.get(user=request.user)
#     school_name=teacher.school
#     query = request.GET.get('q')
#     students1 = user_profile_student.objects.filter(grade=grade)
#     students=students1.filter(school__icontains=school_name).order_by('first_name')

#     if query:
#         students = students.filter(
#             Q(first_name__icontains=query) | Q(last_name__icontains=query) 
#         )
    
#     context = {
#         'grade': grade,
#         'students': students,
#         "school":teacher,
#         'search_query': query
#     }
    
#     return render(request, 'teacher_template/student_report_gradewise.html', context)
def student_report(request):
    user = request.user
    teacher_obj = user_profile_teacher.objects.get(user=user)
    school_name = teacher_obj.school

    students = user_profile_student.objects.filter(school__icontains=school_name).order_by('first_name')

    if school_name in ["Sparsh Global School,Greater Noida", "JP International School,Greater Noida", "SPS,Sonipat", "Satya Prakash Public School, Jabalpur"]:
        valid_grades = range(1, 10)
    elif school_name in ["Vivekanand School,Anand Vihar"]:
        valid_grades = range(6, 10)
    elif school_name in ["Blooming Dale School,Budaun"]:
        valid_grades = range(3, 9)
    else:
        valid_grades = students.values_list('grade', flat=True).distinct()

    students_sectionwise = (
        students.filter(grade__in=valid_grades)
        .values('grade', 'section')
        .annotate(total_number=Count('user_id'))
        .order_by('grade', 'section')
    )

    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(Q(first_name__icontains=search_query) | Q(last_name__icontains=search_query))
        students_sectionwise = []

    grade_section_data = defaultdict(list)
    for entry in students_sectionwise:
        grade_section_data[entry['grade']].append(entry)

    context = {
        "grade_section_data": dict(grade_section_data),  # Convert to regular dict for template
        "students": students,
        "search_query": search_query,
        "school": school_name
    }

    return render(request, 'teacher_template/view_report.html', context)


def student_report_gradewise(request, grade, section):
    user = request.user
    teacher_obj = user_profile_teacher.objects.get(user=user)
    school_name = teacher_obj.school
    query = request.GET.get('q')
    
    students = user_profile_student.objects.filter(grade=grade, section=section)
    student_schoolwise = students.filter(school__icontains=school_name).order_by('first_name')
    
    if query:
        students = students.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query)  
        )
    
    context = {
        'grade': grade,
        'section': section,
        'students': student_schoolwise,
        'school': school_name,
        'search_query': query,
    }
    
    return render(request, 'teacher_template/student_report_gradewise.html', context)


    
def student_detail_report(request,user_id):
    try:
        teacher=user_profile_teacher.objects.get(user=request.user)
        users=user_profile_student.objects.get(user_id=user_id)
        user=User.objects.get(username=users)
        marks=Result.objects.filter(user=user)

        return render(request, "teacher_template/details_mark.html",{"users":users,"marks":marks,"school":teacher})
    except ObjectDoesNotExist:
        # Handle the case when a matching object doesn't exist
        return HttpResponse("The requested data does not exist.")


def teacher_profile(request):
    user = User.objects.get(id=request.user.id)
    school = user_profile_teacher.objects.get(user=user)

    context={
        "user": user,
        "school": school
    }
    return render(request, 'teacher_template/teacher_profile.html', context)

def teacher_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('users:teacher_profile')
    else:
        first_name=request.POST.get('first_name')
        last_name=request.POST.get('last_name')
        school_name = request.POST.get('school_name')
        password = request.POST.get('password')
        mobile= request.POST.get('mobile')
        grade=request.POST.get('grade')
        # profile_pic = request.FILES.get('logo')

        try:
            customuser = User.objects.get(id=request.user.id)
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            teacher = user_profile_teacher.objects.get(user=customuser.id)
            teacher.school=school_name
            teacher.mobile=mobile
            teacher.first_name=first_name
            teacher.last_name=last_name
            teacher.grade=grade
            # school.profile_pic=profile_pic
            teacher.save()
            messages.success(request, "Profile Updated Successfully")
            return redirect('users:teacher_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('users:teacher_profile')
            
def upload_lesson_plan(request):
    if request.method == 'POST':
        form = MacroplannerForm(request.POST, request.FILES)
        if form.is_valid():
            macroplanner_instance = form.save(commit=False)
            macroplanner_instance.user = request.user  # Assuming user is logged in
            macroplanner_instance.save()
            return redirect('users:teacher_home')  # Redirect to a success page or wherever you want
    else:
        form = MacroplannerForm()
    return render(request, 'teacher_template/lesson_plan.html', {'form': form})
    
def upload_micro_plan(request):
    if request.method == 'POST':
        form = MicroplannerForm(request.POST, request.FILES)
        if form.is_valid():
            microplanner_instance = form.save(commit=False)
            microplanner_instance.user = request.user  # Assuming user is logged in
            microplanner_instance.save()
            return redirect('users:teacher_home')  # Redirect to a success page or wherever you want
    else:
        form = MicroplannerForm()
    return render(request, 'teacher_template/microplan.html', {'form': form})
    
def view_macroplanner(request):
    teacher=user_profile_teacher.objects.get(user=request.user)
    school_name=teacher.school
    macroplanners = Macroplanner.objects.filter(school__icontains=school_name)
    return render(request, 'teacher_template/view_macroplanner.html', {'macroplanners': macroplanners})
    
def view_microplanner(request):
    teacher=user_profile_teacher.objects.get(user=request.user)
    school_name=teacher.school
    microplanners = Microplanner.objects.filter(school__icontains=school_name)
    return render(request, 'teacher_template/view_microplanner.html', {'microplanners': microplanners})
    
def advocacy_visit(request):
    teacher_obj = user_profile_teacher.objects.get(user=request.user)
    
    school=teacher_obj.school
    advocacy=AdvocacyVisit.objects.filter(school__icontains=school)
    return render(request, 'teacher_template/advocacy_list.html', {"advocacy":advocacy})
    
def get_report(request, id):
    advocacy_visit = get_object_or_404(AdvocacyVisit, pk=id)
    return render(request, 'teacher_template/advocacy_report.html', {'advocacy_visit': [advocacy_visit]})
    
def download_macroplanner_template(request):
    # Create an empty Excel file template
    wb = openpyxl.Workbook()
    ws = wb.active

    # Populate headers (if any)
    # Example:
    ws['A1'] = 'Month'
    ws['B1'] = 'Name of Subject'
    ws['C1'] = 'No of Session'
    ws['D1'] = 'Assessment'
    ws['E1'] = 'Extra Activities'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="macroplanner_template.xlsx"'

    wb.save(response)
    return response

def download_microplanner_template(request):
    # Create an empty Excel file template
    wb = openpyxl.Workbook()
    ws = wb.active

    # Populate headers (if any)
    # Example:
    ws['A1'] = 'Month'
    ws['B1'] = 'Number of Session'
    ws['C1'] = 'Name of Activity'
    ws['D1'] = 'Extra Activities'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="microplanner_template.xlsx"'

    wb.save(response)
    return response
    
def upload_timetable(request):
    if request.method == 'POST':
        form = TimetableForm(request.POST, request.FILES)
        if form.is_valid():
            timetable_instance = form.save(commit=False)
            timetable_instance.user = request.user # Assuming user is logged in
            timetable_instance.save()
            return redirect('users:teacher_home') # Redirect to a success page or wherever you want
    else:
        form = TimetableForm()
    return render(request, 'teacher_template/timetable.html', {'form': form})

def view_timetable(request):
    teacher=user_profile_teacher.objects.get(user=request.user)
    school_name=teacher.school
    timetable = TimeTable.objects.filter(school__icontains=school_name).order_by('-id').first()
    return render(request, 'teacher_template/view_timetable.html', {'timetable': timetable})
    
def kreativityshow(request):
    school=user_profile_teacher.objects.get(user=request.user)
    school_name=school.school
    kreativityshow=KreativityShow.objects.filter(school__icontains=school_name)
    return render(request, 'teacher_template/kreativityshow.html',{"kreativityshow":kreativityshow})

def guestsession(request):
    school=user_profile_teacher.objects.get(user=request.user)
    school_name=school.school
    guestsession=GuestSession.objects.filter(school__icontains=school_name)
    return render(request, 'teacher_template/guestsession.html',{"guestsession":guestsession})

# def competition(request):
#     school=user_profile_teacher.objects.get(user=request.user)
#     school_name=school.school
#     competition=Competition.objects.filter(school__icontains=school_name)
#     return render(request, 'teacher_template/competition.html',{"competition":competition})
def competition(request):
    school=user_profile_teacher.objects.get(user=request.user)
    school_name=school.school
    competition=Competition.objects.filter(school__icontains=school_name)

    competition_by_date={}
    for comp in competition:
        date=comp.date
        if date in competition_by_date:
            competition_by_date[date].append(comp)
        else:
            competition_by_date[date]=[comp]
    context = {
        'competition_by_date': competition_by_date,
        'user': request.user
    }
        
    return render(request, 'teacher_template/competition.html',context)

# def innovation(request):
#     school=user_profile_teacher.objects.get(user=request.user)
#     school_name=school.school
#     innovations=InnovationClub.objects.filter(school__icontains=school_name)
#     return render(request, 'teacher_template/innovation.html',{"innovations":innovations})
def innovation(request):
    teacher_profile = user_profile_teacher.objects.get(user=request.user)
    school_name = teacher_profile.school
    innovations = InnovationClub.objects.filter(school__icontains=school_name)

    # Group innovations by date
    innovations_by_date = {}
    for innovation in innovations:
        date = innovation.date
        if date in innovations_by_date:
            innovations_by_date[date].append(innovation)
        else:
            innovations_by_date[date] = [innovation]

    context = {
        'innovations_by_date': innovations_by_date,
        'user': request.user
    }

    return render(request, 'teacher_template/innovation.html', context)

def inventory(request):
    school=user_profile_teacher.objects.get(user=request.user)
    school_name=school.school
    inventory_list = Inventory.objects.filter(school__icontains=school_name)
    return render(request, 'teacher_template/inventory.html', {'inventory_list': inventory_list})
    
# def inventory_create_view(request):
#     if request.method == 'POST':
#         inventory_form = InventoryForm(request.POST)
#         kit_formset = KitFormSet(request.POST)

#         if inventory_form.is_valid() and kit_formset.is_valid():
#             inventory = inventory_form.save(commit=False)
#             kits = []
#             for form in kit_formset:
#                 if form.cleaned_data:
#                     kit_name = form.cleaned_data.get('kit_name')
#                     quantity = form.cleaned_data.get('quantity')
#                     kits.append({'kit_name': kit_name, 'quantity': quantity})
#             inventory.kits = kits
#             inventory.save()

#             return redirect('users:teacher_home')
#         else:
#             print("Inventory form errors:", inventory_form.errors)
#             print("Kit formset errors:", kit_formset.errors)
#     else:
#         inventory_form = InventoryForm()
#         kit_formset = KitFormSet()
    
#     return render(request, 'teacher_template/inventory_form.html', {
#         'inventory_form': inventory_form,
#         'kit_formset': kit_formset,
#     })
def inventory_create_view(request):
    if request.method == 'POST':
        inventory_form = InventoryForm(request.POST, request.FILES)

        if inventory_form.is_valid():
            inventory = inventory_form.save(commit=False)
            inventory.kits = []  # or any other appropriate default value
            inventory.save()
            return redirect('users:teacher_home')
        else:
            print("Inventory form errors:", inventory_form.errors)
    else:
        inventory_form = InventoryForm()
    
    return render(request, 'teacher_template/inventory_form.html', {
        'inventory_form': inventory_form,
    })
    
# def competition_create_view(request):
#     if request.method == 'POST':
#         form = CompetitionForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('users:teacher_home') 
#     else:
#         form = CompetitionForm()
    
#     return render(request, 'teacher_template/competition_form.html', {'form': form})
def competition_create_view(request):
    if request.method == 'POST':
        form = CompetitionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users:teacher_home') 
    else:
        form = CompetitionForm()
    
    return render(request, 'teacher_template/competition_form.html', {'form': form})

def load_sections(request):
    school = request.GET.get('school')
    grade = request.GET.get('grade')
    sections = user_profile_student.objects.filter(
        school=school, grade=grade
    ).values_list('section', flat=True).distinct()
    return JsonResponse(list(sections), safe=False)

def load_students(request):
    school = request.GET.get('school')
    grade = request.GET.get('grade')
    section = request.GET.get('section')
    students = user_profile_student.objects.filter(
        school=school, grade=grade, section=section
    ).order_by('first_name')
    return JsonResponse(list(students.values('user_id', 'first_name', 'middle_name', 'last_name')), safe=False)

# def innovationclub_create_view(request):
#     if request.method == 'POST':
#         form = InnovationClubForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('users:teacher_home') 
#     else:
#         form = InnovationClubForm()
    
#     return render(request, 'teacher_template/innovation_form.html', {'form': form})
def innovationclub_create_view(request):
    if request.method == 'POST':
        form = InnovationClubForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users:teacher_home')
    else:
        form = InnovationClubForm()

    return render(request, 'teacher_template/innovation_form.html', {'form': form})
    
def kreativityshow_create_view(request):
    if request.method == 'POST':
        form =KreativityShowForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users:teacher_home') 
    else:
        form = KreativityShowForm()

    return render(request, 'teacher_template/kreativityshow_form.html', {'form': form})
    
    
def guestsession_create_view(request):
    if request.method == 'POST':
        form =GuestSessionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users:teacher_home') 
    else:
        form =GuestSessionForm()
    
    return render
    
def student_login_activity(request):
    user = request.user
    try:
        teacher_profile = user_profile_teacher.objects.get(user=user)
    except user_profile_school.DoesNotExist:
        return render(request, "school/no_permission.html", {"message": "You do not have permission to view this page."})

    form = StudentFilterForm(request.GET or None)
    student_profiles = user_profile_student.objects.filter(school=teacher_profile.school)
    
    if form.is_valid():
        grade = form.cleaned_data.get('grade')
        section = form.cleaned_data.get('section')
        
        if grade:
            student_profiles = student_profiles.filter(grade=grade)
        
        if section:
            student_profiles = student_profiles.filter(section=section)

    login_usernames = [student.user.username for student in student_profiles]
    # students_count=students_count = student_profiles.distinct().count()
    
    # Subquery to get the latest login datetime for each user
    latest_login_subquery = UserLoginActivity.objects.filter(login_username=OuterRef('login_username')).order_by('-login_datetime').values('login_datetime')[:1]
    
    # Get the latest login activity for each user
    latest_login_activities = UserLoginActivity.objects.filter(
        login_username__in=login_usernames,
        login_datetime=Subquery(latest_login_subquery)
    ).distinct().order_by('-login_datetime') 

    # Include user ID in the context
    activities_with_user_id = [
        {
            'login_activity': activity,
            'user_id': User.objects.get(username=activity.login_username).id
        }
        for activity in latest_login_activities
    ]

    context = {
        "activities_with_user_id": activities_with_user_id,
        "school_profile": teacher_profile,
        "form": form,
        # "students_count":students_count
    }

    return render(request, "teacher_template/school_login_activity.html", context)

def student_activity_view(request, user_id):
    activities = UserActivity1.objects.filter(user=user_id).order_by('-date')
    context = {
        'activities': activities,
        'user_id': user_id
    }
    return render(request, 'teacher_template/student_activity.html', context)