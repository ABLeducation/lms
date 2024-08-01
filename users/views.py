from django.shortcuts import render,redirect
from django.contrib.auth import authenticate,logout,login
from django.http import HttpResponse,HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from users.models import Contact,User,Enquiry
from users.forms import *
from django.views.generic import CreateView
from django.contrib import messages
from users.Emailbackend import *
import requests
import json
from .utils import activateEmail
import openpyxl

# Create your views here.
def index(request):
    schools = School.objects.all()
    grouped_schools = [schools[i:i+4] for i in range(0, len(schools), 4)]
    
    allowed_emails = [
        'training@thinnkware.com',
        'csahu@thinnkware.com',
        'training3@thinnkware.com',
        'training2@thinnkware.com',
        'training10@thinnkware.com',
        'training13@thinnkware.com',
        'developer2@thinnkware.com',
        'manager.lnd@thinnkware.com'
    ]
    context = {
        'grouped_schools': grouped_schools,
        'allowed_emails':allowed_emails
    }
    return render(request,"users/index.html", context)

class StudentSignUpView(CreateView):
    model = User
    form_class = studentsignupform
    template_name = 'users/student.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'student'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        to=user.email
        activateEmail(self.request, user, to)
        return redirect('users:message')

class ParentSignUpView(CreateView):
    model = User
    form_class = parentsignupform
    template_name = 'users/parent.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'parent'
        return super().get_context_data(**kwargs)

    def form_valid(self,form):
        user = form.save()
        to=user.email
        activateEmail(self.request, user, to) 
        return redirect('users:message')

class TeacherSignUpView(CreateView):
    model = User
    form_class = teachersignupform
    template_name = 'users/teacher.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'teacher'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        to=user.email
        activateEmail(self.request, user, to)
        return redirect('users:message')

class PrincipalSignUpView(CreateView):
    model = User
    form_class = principalsignupform
    template_name = 'users/principal.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'principal'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        to=user.email
        activateEmail(self.request, user, to)
        return redirect('users:message')

class SchoolSignUpView(CreateView):
    model = User
    form_class = schoolsignupform
    template_name = 'users/school.html'

    def get_context_data(self, **kwargs):
        kwargs['user_type'] = 'school'
        return super().get_context_data(**kwargs)

    def form_valid(self, form):
        user = form.save()
        to=user.email
        activateEmail(self.request, user, to)
        return redirect('users:message')
        
def register_student_from_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid Excel file (.xlsx)')
            return redirect('register_student_from_excel')
        
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        success_count = 0
        error_count = 0
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Assuming data starts from the second row
            try:
                username,first_name, last_name, email, password, dob, grade, section, school, country, state, city = row
                student_data = {
                    'username': username,
                    'email': email,
                    'password1': password,
                    'password2': password,
                    'First_Name': first_name,
                    'Middle_Name': None,
                    'Last_Name': last_name,
                    'dob': dob,
                    'grade': grade,
                    'section': section,
                    'school': school,
                    'country': country,
                    'state': state,
                    'city': city
                }
                form = studentsignupform(student_data)
                if form.is_valid():
                    form.save()
                    success_count += 1
                else:
                    error_count += 1
                    error_msg = f"Error in row {row_index}: {form.errors.as_text()}"
                    messages.error(request, error_msg)
            except Exception as e:
                error_count += 1
                error_msg = f"Error processing row {row_index}: {e}"
                messages.error(request, error_msg)
                print(error_msg)
        
        messages.success(request, f'{success_count} students registered successfully from Excel file.')
        if error_count > 0:
            messages.error(request, f'Failed to register {error_count} students from Excel file.')
        return redirect('users:index')  # Redirect to a success page or any other page

    return render(request, 'users/registration_with_excel.html')

def download_excel_template(request):
    # Create an empty Excel file template
    wb = openpyxl.Workbook()
    ws = wb.active

    # Populate headers (if any)
    # Example:
    ws['A1'] = 'username'
    ws['B1'] = 'first_name'
    ws['C1'] = 'last_name'
    ws['D1'] = 'email'
    ws['E1'] = 'password'
    ws['F1'] = 'dob'
    ws['G1'] = 'grade'
    ws['H1'] = 'section'
    ws['I1'] = 'school'
    ws['J1'] = 'country'
    ws['K1'] = 'state'
    ws['L1'] = 'city'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_template.xlsx"'

    wb.save(response)
    return response

def user_login(request):
    if request.method == "POST":
        # email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        capcha_token=request.POST.get("g-recaptcha-response")
        cap_url="https://www.google.com/recaptcha/api/siteverify"
        cap_secret="6LeLgcEkAAAAAPQMUQSoVzHMQwwmCQx_UATRgoaE"
        cap_data={"secret":cap_secret, "response":capcha_token}
        cap_server_response=requests.post(url=cap_url, data=cap_data)
        cap_json=json.loads(cap_server_response.text)
        if cap_json['success']==False:
            messages.error(request,"Invalid Captcha Try Again")
            return HttpResponseRedirect("/")
            
        user = EmailOrUsernameBackend().authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                # Set the backend attribute on the user
                user.backend = 'users.Emailbackend.EmailOrUsernameBackend'
                login(request, user)
                if user_profile_school.objects.filter(user=user).exists():
                    return redirect('users:school_home')
                elif user_profile_teacher.objects.filter(user=user).exists():
                    return redirect('users:teacher_home')
                else:
                    return redirect('users:index')
            else:
                return HttpResponse("User not verified!")
        else:
            return HttpResponse("Please use correct id and password")

    else:
        return render(request, 'users/login.html')

@login_required
def user_logout(request):
    logout(request)
    return render(request, 'users/index.html')

def register(request):
    return render(request, 'users/register.html')

def contact(request):
    if request.method=="POST":
        name=request.POST.get('name')
        contact_num=request.POST.get('contact')
        email=request.POST.get('email')
        message=request.POST.get('message')

        con=Contact(name=name, contact_no=contact_num, mail=email, message=message)
        con.save()
    return render(request, "users/contact.html")  
    
def editor(request):
    return render(request, "users/editor.html")
    
def enquiry(request):
    if request.method=="POST":
        name=request.POST.get('name')
        contact_num=request.POST.get('contact')
        email=request.POST.get('email')
        enquiry=request.POST.get('query')
        
        capcha_token=request.POST.get("g-recaptcha-response")
        cap_url="https://www.google.com/recaptcha/api/siteverify"
        cap_secret="6LeLgcEkAAAAAPQMUQSoVzHMQwwmCQx_UATRgoaE"
        cap_data={"secret":cap_secret, "response":capcha_token}
        cap_server_response=requests.post(url=cap_url, data=cap_data)
        print(cap_server_response.text)
        cap_json=json.loads(cap_server_response.text)
        if cap_json['success']==False:
            messages.error(request,"Invalid Captcha Try Again")
            return HttpResponseRedirect("/")

        enquiry1=Enquiry(name=name, contact=contact_num, email=email, query=enquiry)
        enquiry1.save()

    return HttpResponse("Thank you for your response. We will get back to you.")
    
def message(request):
    return render(request,"users/message.html")

def create_student_project(request):
    if request.method == 'POST':
        form = StudentProjectForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('users:success_page')
    else:
        form = StudentProjectForm()
    return render(request, 'users/student_project_form.html', {'form': form})
    
def success_page(request):
    return render(request,'users/demo_success.html')
    
def create_advocacy_visit(request):
    if request.method == 'POST':
        form = AdvocacyVisitForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('users:index')
    else:
        form = AdvocacyVisitForm()
    return render(request, 'users/advocay.html', {'form': form})